import json
from pprint import pprint

import openpyxl


class ExcelHandler:
    def __init__(self, file_path):
        self.file = file_path
        self.wb = None

    def open_excel(self):
        # 打开表格
        wb = openpyxl.load_workbook(self.file)
        self.wb = wb
        return wb

    def get_sheet(self, name):
        wb = self.open_excel()
        # 获取工作簿
        sheet = wb[name]
        return sheet

    def read_sheet(self, name):
        # 读取某个工作簿
        sheet = self.get_sheet(name)
        # 获取表格内所有的数据，然后转化为list
        data = list(sheet.rows)
        # 表头
        headers = []
        case_list = []
        for header in data[0]:
            headers.append(header.value)

        for list1 in data[1:]:
            # print(list1)
            case_dict = {}
            for idx, cell in enumerate(list1):
                case_dict[headers[idx]] = cell.value
            case_list.append(case_dict)
        self.close()
        return case_list

    def write(self, sheet_name, row, cloumn, data):
        # 得到表单，修改表单
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row, cloumn).value = data
        self.save()
        self.close()

    def write1(self, sheet_name, row, cloumn, key,  data):
        sheet = self.get_sheet(sheet_name)
        json.loads(sheet.cell(row, cloumn).value)[key] = data
        self.save()
        self.close()

    def save(self):
        self.wb.save(self.file)

    def close(self):
        self.wb.close()



if __name__ == '__main__':
    workbook = ExcelHandler("../testdata/cases.xlsx")
    # print(workbook.get_sheet("Sheet1"))
    pprint(workbook.read_sheet("Sheet2"))
    # print(workbook.write("Sheet1", 2, 1, 1))